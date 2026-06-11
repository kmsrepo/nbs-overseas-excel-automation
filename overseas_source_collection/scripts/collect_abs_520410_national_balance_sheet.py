#!/usr/bin/env python3
"""Collect latest ABS 5204.0 Table 10 National Balance Sheet and derive net financial worth."""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import pathlib
import re
import subprocess
from urllib.parse import urljoin

from openpyxl import load_workbook


ROOT = pathlib.Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "abs_520410_national_balance_sheet"
RAW_DIR = OUT_DIR / "raw"
LATEST_RELEASE_URL = (
    "https://www.abs.gov.au/statistics/economy/national-accounts/"
    "australian-system-national-accounts/latest-release"
)
START_YEAR = 2007
END_YEAR = 2025
FINANCIAL_ASSETS_SERIES_ID = "A2421138R"
LIABILITIES_SERIES_ID = "A2421145L"
WORKBOOK_NAME = "5204010_National_Balance_Sheet.xlsx"
RAW_WORKBOOK = RAW_DIR / WORKBOOK_NAME
LONG_CSV = OUT_DIR / "abs_520410_순금융자산_2007_2025.csv"
LOG_CSV = OUT_DIR / "abs_520410_수집로그.csv"


def fetch_bytes(url: str, accept: str | None = None) -> bytes:
    command = ["curl", "-L", "--fail", "--silent", "--show-error", "--max-time", "180"]
    command += ["--retry", "3", "--retry-delay", "2"]
    command += ["-A", "Mozilla/5.0 Codex ABS collector"]
    if accept:
        command += ["-H", f"Accept: {accept}"]
    command.append(url)
    result = subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if not result.stdout.strip():
        raise RuntimeError(f"empty response: {url}")
    return result.stdout


def latest_workbook_url() -> str:
    html = fetch_bytes(LATEST_RELEASE_URL, "text/html").decode("utf-8", errors="replace")
    match = re.search(r'href="([^"]*5204010_National_Balance_Sheet\.xlsx[^"]*)"', html)
    if not match:
        raise RuntimeError("ABS latest release page did not contain 5204010_National_Balance_Sheet.xlsx")
    return urljoin(LATEST_RELEASE_URL, match.group(1))


def write_csv(path: pathlib.Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def series_columns(workbook_path: pathlib.Path) -> dict[str, int]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["Data1"]
    columns: dict[str, int] = {}
    for col_no in range(2, ws.max_column + 1):
        series_id = ws.cell(row=10, column=col_no).value
        if series_id in {FINANCIAL_ASSETS_SERIES_ID, LIABILITIES_SERIES_ID}:
            columns[str(series_id)] = col_no
    missing = {FINANCIAL_ASSETS_SERIES_ID, LIABILITIES_SERIES_ID} - set(columns)
    if missing:
        raise RuntimeError(f"ABS workbook missing required series: {sorted(missing)}")
    return columns


def derive_rows(workbook_path: pathlib.Path, collected_at: str, source_url: str, digest: str) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["Data1"]
    columns = series_columns(workbook_path)
    rows: list[dict[str, str]] = []
    for row_no in range(12, ws.max_row + 1):
        period = ws.cell(row=row_no, column=1).value
        year = getattr(period, "year", None)
        if year is None or not (START_YEAR <= year <= END_YEAR):
            continue
        financial_assets = float(ws.cell(row=row_no, column=columns[FINANCIAL_ASSETS_SERIES_ID]).value)
        liabilities = float(ws.cell(row=row_no, column=columns[LIABILITIES_SERIES_ID]).value)
        net_financial_worth = round(financial_assets - liabilities, 1)
        rows.append(
            {
                "수집일시": collected_at,
                "출처기관": "Australian Bureau of Statistics",
                "자료명": "5204.0 Australian System of National Accounts Table 10 National Balance Sheet",
                "국가코드": "AUS",
                "국가명": "Australia",
                "연도": str(year),
                "기준시점": period.strftime("%Y-%m-%d"),
                "금융자산_SeriesID": FINANCIAL_ASSETS_SERIES_ID,
                "부채_SeriesID": LIABILITIES_SERIES_ID,
                "금융자산_십억AUD": f"{financial_assets:.1f}",
                "부채_십억AUD": f"{liabilities:.1f}",
                "순금융자산_십억AUD": f"{net_financial_worth:.1f}",
                "순금융자산_백만AUD": str(int(round(net_financial_worth * 1000))),
                "단위": "백만 AUD",
                "계산식": f"{FINANCIAL_ASSETS_SERIES_ID} - {LIABILITIES_SERIES_ID}",
                "원자료주소": source_url,
                "최신릴리스주소": LATEST_RELEASE_URL,
                "원본파일": str(workbook_path.relative_to(ROOT)),
                "sha256": digest,
            }
        )
    return rows


def write_readme(source_url: str) -> None:
    (OUT_DIR / "README.md").write_text(
        "\n".join(
            [
                "# ABS 5204.0 Table 10 National Balance Sheet",
                "",
                "Australian Bureau of Statistics 최신 릴리스에서 5204010 National Balance Sheet 파일을 수집합니다.",
                "",
                f"- 최신 릴리스 URL: {LATEST_RELEASE_URL}",
                f"- 원본 xlsx URL: {source_url}",
                f"- 원본 파일: raw/{WORKBOOK_NAME}",
                f"- 순금융자산 계산식: {FINANCIAL_ASSETS_SERIES_ID} - {LIABILITIES_SERIES_ID}",
                "- 단위 변환: billion AUD * 1000 = million AUD",
                "",
            ]
        ),
        encoding="utf-8",
    )


def main() -> int:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    collected_at = dt.datetime.now().astimezone().isoformat(timespec="seconds")
    source_url = latest_workbook_url()
    content = fetch_bytes(source_url, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    RAW_WORKBOOK.write_bytes(content)
    digest = hashlib.sha256(content).hexdigest()

    rows = derive_rows(RAW_WORKBOOK, collected_at, source_url, digest)
    write_csv(
        LONG_CSV,
        [
            "수집일시",
            "출처기관",
            "자료명",
            "국가코드",
            "국가명",
            "연도",
            "기준시점",
            "금융자산_SeriesID",
            "부채_SeriesID",
            "금융자산_십억AUD",
            "부채_십억AUD",
            "순금융자산_십억AUD",
            "순금융자산_백만AUD",
            "단위",
            "계산식",
            "원자료주소",
            "최신릴리스주소",
            "원본파일",
            "sha256",
        ],
        rows,
    )
    write_csv(
        LOG_CSV,
        [
            "수집일시",
            "출처기관",
            "자료명",
            "요청기간",
            "원자료주소",
            "최신릴리스주소",
            "원본파일",
            "파일크기_bytes",
            "sha256",
            "상태",
        ],
        [
            {
                "수집일시": collected_at,
                "출처기관": "Australian Bureau of Statistics",
                "자료명": "5204010 National Balance Sheet",
                "요청기간": f"{START_YEAR}-{END_YEAR}",
                "원자료주소": source_url,
                "최신릴리스주소": LATEST_RELEASE_URL,
                "원본파일": str(RAW_WORKBOOK.relative_to(ROOT)),
                "파일크기_bytes": str(RAW_WORKBOOK.stat().st_size),
                "sha256": digest,
                "상태": "완료",
            }
        ],
    )
    write_readme(source_url)
    print(f"rows={len(rows)}")
    print(f"source_url={source_url}")
    print(f"workbook={RAW_WORKBOOK}")
    print(f"csv={LONG_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
