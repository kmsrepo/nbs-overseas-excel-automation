#!/usr/bin/env python3
"""GitHub Actions friendly collector and requested pivot workbook builder."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import pathlib
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = pathlib.Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "outputs" / "requested_pivot_workbook"
FINAL_WORKBOOK = OUT_DIR / "요청사항반영_피벗통합_메모포함.xlsx"
BUILD_LOG = OUT_DIR / "요청사항반영_엑셀생성로그.csv"
SOURCE_ARCHIVE = OUT_DIR / "요청사항반영_원본자료.zip"

NORMALIZER_SCRIPTS = [
    ROOT / "overseas_source_collection/scripts/collect_fred_us_series.py",
    ROOT / "overseas_source_collection/scripts/collect_oecd_table9b_nonfinancial_assets.py",
    ROOT / "overseas_source_collection/scripts/collect_oecd_financial_net_worth.py",
    ROOT / "overseas_source_collection/scripts/collect_abs_520410_national_balance_sheet.py",
    ROOT / "overseas_source_collection/scripts/collect_oecd_gdp_output_expenditure.py",
]

PAYLOAD_SCRIPTS = [
    ROOT / "outputs/fred_us_pivot/build_fred_pivot_payload.py",
    ROOT / "outputs/oecd_table9b_pivot/build_oecd_pivot_payload.py",
    ROOT / "outputs/oecd_financial_net_worth_pivot/build_financial_net_worth_payload.py",
    ROOT / "outputs/oecd_gdp_pivot/build_gdp_pivot_payload.py",
    ROOT / "outputs/oecd_gdp_pivot/build_gdp_oecd_members_payload.py",
]

PAYLOAD_PATHS = {
    "fred": ROOT / "outputs/fred_us_pivot/fred_pivot_payload.json",
    "table9b": ROOT / "outputs/oecd_table9b_pivot/oecd_pivot_payload.json",
    "financial": ROOT / "outputs/oecd_financial_net_worth_pivot/financial_net_worth_pivot_payload.json",
    "gdp": ROOT / "outputs/oecd_gdp_pivot/gdp_oecd_members_pivot_payload.json",
}

SOURCE_ARCHIVE_DIRS = [
    ROOT / "overseas_source_collection/fred_us",
    ROOT / "overseas_source_collection/abs_520410_national_balance_sheet",
    ROOT / "overseas_source_collection/oecd_table9b_nonfinancial_assets",
    ROOT / "overseas_source_collection/oecd_financial_net_worth",
    ROOT / "overseas_source_collection/oecd_gdp_output_expenditure",
]
SOURCE_ARCHIVE_FILES = [
    ROOT / "overseas_source_collection/source_targets.csv",
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BASE_FONT = Font(name="Calibri", size=10)
THIN_GRAY = Side(style="thin", color="D9E2F3")
PURPLE_FILL = PatternFill("solid", fgColor="D9D2E9")
PURPLE_FONT = Font(name="Calibri", size=10, bold=True, color="351C75")
PURPLE_BORDER = Border(outline=True, left=Side(style="thin", color="674EA7"), right=Side(style="thin", color="674EA7"), top=Side(style="thin", color="674EA7"), bottom=Side(style="thin", color="674EA7"))
KOREA_FILL = PatternFill("solid", fgColor="E2F0D9")
KOREA_FONT_COLOR = "0000FF"
ABS_DERIVED_FILL = PatternFill("solid", fgColor="EADCF8")
FRED_US_NONFINANCIAL_FILL = PatternFill("solid", fgColor="F4CCCC")
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

ET.register_namespace("", SPREADSHEET_NS)
ET.register_namespace("r", OFFICE_REL_NS)


def write_csv(path: pathlib.Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def run_step(label: str, command: list[str]) -> dict[str, str]:
    print(f"[{label}] {' '.join(command)}", flush=True)
    started = dt.datetime.now().astimezone()
    result = subprocess.run(command, cwd=ROOT, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    finished = dt.datetime.now().astimezone()
    row = {
        "단계": label,
        "명령": " ".join(command),
        "시작일시": started.isoformat(timespec="seconds"),
        "종료일시": finished.isoformat(timespec="seconds"),
        "상태": "완료" if result.returncode == 0 else "실패",
        "종료코드": str(result.returncode),
        "표준출력": result.stdout.strip(),
        "표준오류": result.stderr.strip(),
    }
    if result.stdout.strip():
        print(result.stdout.strip(), flush=True)
    if result.returncode != 0:
        if result.stderr.strip():
            print(result.stderr.strip(), flush=True)
        raise RuntimeError(f"{label} failed")
    return row


def include_in_source_archive(path: pathlib.Path) -> bool:
    if path.name == ".DS_Store" or path.suffix == ".pyc":
        return False
    return "__pycache__" not in path.parts


def create_source_archive() -> int:
    SOURCE_ARCHIVE.parent.mkdir(parents=True, exist_ok=True)
    SOURCE_ARCHIVE.unlink(missing_ok=True)
    file_count = 0
    with zipfile.ZipFile(SOURCE_ARCHIVE, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for source_dir in SOURCE_ARCHIVE_DIRS:
            if not source_dir.exists():
                continue
            for path in sorted(source_dir.rglob("*")):
                if not path.is_file() or not include_in_source_archive(path):
                    continue
                archive.write(path, path.relative_to(ROOT).as_posix())
                file_count += 1
        for source_file in SOURCE_ARCHIVE_FILES:
            if not source_file.exists() or not source_file.is_file() or not include_in_source_archive(source_file):
                continue
            archive.write(source_file, source_file.relative_to(ROOT).as_posix())
            file_count += 1
    return file_count


def load_payloads() -> dict[str, Any]:
    return {name: json.loads(path.read_text(encoding="utf-8")) for name, path in PAYLOAD_PATHS.items()}


def append_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def format_sheet(ws, freeze_cols: int = 0) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    ws.freeze_panes = ws.cell(row=2, column=freeze_cols + 1).coordinate if freeze_cols else "A2"
    border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = BASE_FONT
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter][: min(ws.max_row, 200)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = max(10, min(45, max_len + 2))


def set_number_format(ws, start_col: int, start_row: int, number_format: str) -> None:
    for row in ws.iter_rows(min_row=start_row, min_col=start_col, max_col=ws.max_column):
        for cell in row:
            cell.number_format = number_format


def style_fallback_cell(cell, number_format: str = "#,##0") -> None:
    cell.fill = PURPLE_FILL
    cell.font = PURPLE_FONT
    cell.border = PURPLE_BORDER
    cell.number_format = number_format


def format_comment_cell(cell, number_format: str = "#,##0") -> None:
    cell.number_format = number_format


def style_korea_rows(ws, country_col: int) -> int:
    styled_rows = 0
    for row_no in range(2, ws.max_row + 1):
        if ws.cell(row=row_no, column=country_col).value != "Korea":
            continue
        styled_rows += 1
        for col_no in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_no, column=col_no)
            cell.fill = KOREA_FILL
            cell.font = Font(
                name=cell.font.name or "Calibri",
                size=cell.font.sz or 10,
                bold=cell.font.bold,
                italic=cell.font.italic,
                underline=cell.font.underline,
                strike=cell.font.strike,
                color=KOREA_FONT_COLOR,
            )
    return styled_rows


def style_abs_derived_rows(ws, code_col: int) -> int:
    styled_rows = 0
    for row_no in range(2, ws.max_row + 1):
        if ws.cell(row=row_no, column=code_col).value != "AUS_ABS":
            continue
        styled_rows += 1
        for col_no in range(1, ws.max_column + 1):
            ws.cell(row=row_no, column=col_no).fill = ABS_DERIVED_FILL
    return styled_rows


def style_fred_us_derived_rows(ws, row_numbers: list[int]) -> int:
    for row_no in row_numbers:
        for col_no in range(1, ws.max_column + 1):
            ws.cell(row=row_no, column=col_no).fill = FRED_US_NONFINANCIAL_FILL
    return len(row_numbers)


def fred_sheet_indexes(wb: Workbook) -> tuple[dict[str, int], dict[str, int]]:
    fred_ws = wb["FRED_피벗"]
    year_to_col = {str(cell.value): cell.column for cell in fred_ws[1][2:] if cell.value is not None}
    code_to_row = {
        str(fred_ws.cell(row=row_no, column=1).value): row_no
        for row_no in range(2, fred_ws.max_row + 1)
        if fred_ws.cell(row=row_no, column=1).value
    }
    return code_to_row, year_to_col


def fred_sum_formula(code_to_row: dict[str, int], year_to_col: dict[str, int], year: str, codes: list[str]) -> str:
    missing_codes = [code for code in codes if code not in code_to_row]
    if missing_codes:
        raise KeyError(f"FRED_피벗 sheet is missing required code rows: {', '.join(missing_codes)}")
    if year not in year_to_col:
        raise KeyError(f"FRED_피벗 sheet is missing required year column: {year}")

    col_letter = get_column_letter(year_to_col[year])
    row_numbers = [code_to_row[code] for code in codes]
    sorted_rows = sorted(row_numbers)
    if sorted_rows == list(range(sorted_rows[0], sorted_rows[-1] + 1)):
        return f"=SUM('FRED_피벗'!{col_letter}{sorted_rows[0]}:{col_letter}{sorted_rows[-1]})"
    refs = ",".join(f"'FRED_피벗'!{col_letter}{row_no}" for row_no in row_numbers)
    return f"=SUM({refs})"


def fred_sum_cached_value(
    wb: Workbook,
    code_to_row: dict[str, int],
    year_to_col: dict[str, int],
    year: str,
    codes: list[str],
) -> float:
    fred_ws = wb["FRED_피벗"]
    col_no = year_to_col[year]
    values = [fred_ws.cell(row=code_to_row[code], column=col_no).value for code in codes]
    return sum(float(value or 0) for value in values)


def cached_number_text(value: float) -> str:
    if abs(value - round(value)) < 0.0000001:
        return str(int(round(value)))
    return format(value, ".15g")


def sheet_xml_path(archive: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_id = None
    for sheet in workbook_root.findall(f".//{{{SPREADSHEET_NS}}}sheet"):
        if sheet.attrib.get("name") == sheet_name:
            rel_id = sheet.attrib[f"{{{OFFICE_REL_NS}}}id"]
            break
    if rel_id is None:
        raise KeyError(f"Workbook is missing sheet: {sheet_name}")

    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for relationship in rels_root.findall(f".//{{{PACKAGE_REL_NS}}}Relationship"):
        if relationship.attrib.get("Id") != rel_id:
            continue
        target = relationship.attrib["Target"]
        return target.lstrip("/") if target.startswith("/") else f"xl/{target}"
    raise KeyError(f"Workbook is missing relationship for sheet: {sheet_name}")


def add_formula_cached_values(workbook_path: pathlib.Path, sheet_name: str, cached_values: dict[str, float]) -> None:
    if not cached_values:
        return
    with zipfile.ZipFile(workbook_path, "r") as archive:
        target_sheet = sheet_xml_path(archive, sheet_name)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=workbook_path.parent) as temp_file:
        temp_path = pathlib.Path(temp_file.name)

    try:
        with zipfile.ZipFile(workbook_path, "r") as source, zipfile.ZipFile(
            temp_path,
            "w",
            compression=zipfile.ZIP_DEFLATED,
        ) as target:
            for item in source.infolist():
                content = source.read(item.filename)
                if item.filename == target_sheet:
                    content = add_cached_values_to_sheet_xml(content, cached_values)
                target.writestr(item, content)
        temp_path.replace(workbook_path)
    finally:
        temp_path.unlink(missing_ok=True)


def add_cached_values_to_sheet_xml(sheet_xml: bytes, cached_values: dict[str, float]) -> bytes:
    root = ET.fromstring(sheet_xml)
    for coordinate, value in cached_values.items():
        cell = root.find(f".//{{{SPREADSHEET_NS}}}c[@r='{coordinate}']")
        if cell is None or cell.find(f"{{{SPREADSHEET_NS}}}f") is None:
            continue
        cached = cell.find(f"{{{SPREADSHEET_NS}}}v")
        if cached is None:
            cached = ET.Element(f"{{{SPREADSHEET_NS}}}v")
            formula_index = next(
                (index for index, child in enumerate(list(cell)) if child.tag == f"{{{SPREADSHEET_NS}}}f"),
                -1,
            )
            cell.insert(formula_index + 1 if formula_index >= 0 else len(cell), cached)
        cached.text = cached_number_text(value)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def add_comment(cell, text: str) -> None:
    cell.comment = Comment(text, "Codex")


def build_fred_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("FRED_피벗")
    headers = ["코드", "코드항목명", *payload["years"]]
    rows = [
        headers,
        *[
            [row["코드"], row["코드항목명"], *[row.get(year) for year in payload["years"]]]
            for row in payload["pivot_rows"]
        ],
    ]
    append_rows(ws, rows)
    format_sheet(ws, freeze_cols=2)
    set_number_format(ws, 3, 2, "#,##0")


def build_table9b_sheet(wb: Workbook, payload: dict[str, Any]) -> dict[str, float]:
    ws = wb.create_sheet("비금융자산_피벗")
    headers = ["트랜잭션코드", "트랜잭션명", "나라", "화폐단위", *payload["years"]]
    row_index: dict[tuple[str, str, str, str], int] = {}
    fred_us_derived_rows: list[int] = []
    fred_formula_rows: list[tuple[int, list[str]]] = []
    rows = [headers]
    for index, row in enumerate(payload["pivot_rows"], start=2):
        key = (row["트랜잭션코드"], row["트랜잭션명"], row["나라"], row["화폐단위"])
        row_index[key] = index
        if row.get("행강조") in {"FRED_US_NONFINANCIAL_NN", "FRED_US_REAL_ESTATE"}:
            fred_us_derived_rows.append(index)
        if row.get("FRED합산코드"):
            fred_formula_rows.append((index, list(row["FRED합산코드"])))
        rows.append([row["트랜잭션코드"], row["트랜잭션명"], row["나라"], row["화폐단위"], *[row.get(year) for year in payload["years"]]])
    append_rows(ws, rows)
    fred_code_to_row, fred_year_to_col = fred_sheet_indexes(wb)
    formula_cached_values: dict[str, float] = {}
    for row_no, codes in fred_formula_rows:
        for year_idx, year in enumerate(payload["years"], start=5):
            cell = ws.cell(row=row_no, column=year_idx)
            year_text = str(year)
            cell.value = fred_sum_formula(fred_code_to_row, fred_year_to_col, year_text, codes)
            formula_cached_values[cell.coordinate] = fred_sum_cached_value(
                wb,
                fred_code_to_row,
                fred_year_to_col,
                year_text,
                codes,
            )
    format_sheet(ws, freeze_cols=4)
    set_number_format(ws, 5, 2, "#,##0")
    for note in payload["notes"]:
        key = (note["트랜잭션코드"], note["트랜잭션명"], note["나라"], note["화폐단위"])
        row_no = row_index.get(key)
        year_idx = payload["years"].index(note["연도"]) if note["연도"] in payload["years"] else -1
        if not row_no or year_idx < 0:
            continue
        cell = ws.cell(row=row_no, column=5 + year_idx)
        format_comment_cell(cell, "#,##0")
        add_comment(
            cell,
            "\n".join(
                [
                    f"OECD 관측상태: {note['관측상태코드']} - {note['관측상태명']}",
                    f"트랜잭션: {note['트랜잭션코드']} {note['트랜잭션명']}",
                    f"나라/통화: {note['나라']} / {note['화폐단위']}",
                    f"연도: {note['연도']}",
                    f"값(백만 현지통화): {note['값_백만현지통화']}",
                ]
            ),
        )
    style_korea_rows(ws, country_col=3)
    style_fred_us_derived_rows(ws, fred_us_derived_rows)

    code_ws = wb.create_sheet("비금융자산_코드")
    append_rows(code_ws, [["트랜잭션코드", "트랜잭션명"], *[[row["트랜잭션코드"], row["트랜잭션명"]] for row in payload["codes"]]])
    format_sheet(code_ws, freeze_cols=1)
    return formula_cached_values


def build_financial_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("금융순자산_피벗")
    headers = ["나라", "국가코드", "화폐단위", "지표코드", "지표명", "단위", *payload["years"]]
    row_index: dict[tuple[str, str], int] = {}
    rows = [headers]
    for index, row in enumerate(payload["pivot_rows"], start=2):
        row_index[(row["나라"], row["화폐단위"])] = index
        rows.append([row["나라"], row["국가코드"], row["화폐단위"], row["지표코드"], row["지표명"], row["단위"], *[row.get(year) for year in payload["years"]]])
    append_rows(ws, rows)
    format_sheet(ws, freeze_cols=6)
    set_number_format(ws, 7, 2, "#,##0")
    for note in payload["notes"]:
        row_no = row_index.get((note["나라"], note["화폐단위"]))
        year_idx = payload["years"].index(note["연도"]) if note["연도"] in payload["years"] else -1
        if not row_no or year_idx < 0:
            continue
        cell = ws.cell(row=row_no, column=7 + year_idx)
        format_comment_cell(cell, "#,##0")
        add_comment(
            cell,
            "\n".join(
                [
                    f"OECD 관측상태: {note['관측상태코드']} - {note['관측상태명']}",
                    f"지표: {note['지표코드']} / {note['새지표코드']} {note['지표명']}",
                    f"나라/통화: {note['나라']} / {note['화폐단위']}",
                    f"연도: {note['연도']}",
                    f"값(백만 현지통화): {note['값_백만현지통화']}",
                ]
            ),
        )
    style_korea_rows(ws, country_col=1)
    style_abs_derived_rows(ws, code_col=2)


def build_gdp_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("GDP_OECD회원국")
    headers = ["나라", "국가코드", "화폐단위", "단위", *payload["years"]]
    row_index: dict[str, int] = {}
    rows = [headers]
    for index, row in enumerate(payload["pivot_rows"], start=2):
        row_index[row["국가코드"]] = index
        rows.append([row["나라"], row["국가코드"], row["화폐단위"], row["단위"], *[row.get(year) for year in payload["years"]]])
    append_rows(ws, rows)
    format_sheet(ws, freeze_cols=4)
    set_number_format(ws, 5, 2, "#,##0")
    style_korea_rows(ws, country_col=1)
    fallback_addresses: set[tuple[int, int]] = set()
    for fallback in payload["fallback_cells"]:
        row_no = row_index.get(fallback["국가코드"])
        year_idx = payload["years"].index(fallback["연도"]) if fallback["연도"] in payload["years"] else -1
        if not row_no or year_idx < 0:
            continue
        cell = ws.cell(row=row_no, column=5 + year_idx)
        fallback_addresses.add((cell.row, cell.column))
        style_fallback_cell(cell)
    for note in payload["notes"]:
        row_no = row_index.get(note["국가코드"])
        year_idx = payload["years"].index(note["연도"]) if note["연도"] in payload["years"] else -1
        if not row_no or year_idx < 0:
            continue
        cell = ws.cell(row=row_no, column=5 + year_idx)
        if (cell.row, cell.column) in fallback_addresses:
            style_fallback_cell(cell)
        else:
            format_comment_cell(cell, "#,##0")
        add_comment(
            cell,
            "\n".join(
                [
                    f"OECD 관측상태: {note['관측상태코드']} - {note['관측상태명']}",
                    f"접근법: {note['접근법']}",
                    f"거래: {note['거래코드']} {note['거래명']}",
                    f"나라/통화: {note['나라']} / {note['화폐단위']}",
                    f"연도: {note['연도']}",
                    f"값(백만 현지통화): {note['값_백만현지통화']}",
                ]
            ),
        )


def build_info_sheet(wb: Workbook, payloads: dict[str, Any]) -> None:
    ws = wb.create_sheet("수집정보")
    rows = [
        ["항목", "내용"],
        ["FRED", f"행 {payloads['fred']['summary']['행수']}, 연도 {payloads['fred']['years'][0]}-{payloads['fred']['years'][-1]}, 코드순서 유지"],
        ["OECD 비금융자산", f"Total economy, 요청 트랜잭션 순서 유지, 비확정값 메모 {payloads['table9b']['summary']['비확정값_노트수']}개"],
        ["FRED 미국 비금융자산", f"OECD 비금융자산 NN 미국 결측분을 FRED 6개 코드 SUM 수식으로 보강 {payloads['table9b']['summary'].get('FRED_US_NN_보강행수', 0)}행, 부동산 3개 코드 SUM 수식 {payloads['table9b']['summary'].get('FRED_US_부동산_보강행수', 0)}행, 연한 빨간색 표시"],
        ["OECD 금융순자산", f"국가 행/연도 열 피벗, 비확정값 메모 {payloads['financial']['summary']['메모수']}개"],
        ["ABS 호주 순금융자산", f"ABS 520410 산출 행 {payloads['financial']['summary'].get('ABS_520410_보강행수', 0)}개, 금융자산-부채로 계산 후 옅은 보라색 표시"],
        ["OECD GDP", f"OECD 회원국 {payloads['gdp']['summary']['OECD회원국수']}개, 지출접근 보강 {payloads['gdp']['summary']['지출접근보강셀수']}셀, 비확정값 메모 {payloads['gdp']['summary']['메모수']}개"],
        ["GDP 한국 예외", f"한국 산출접근 값이 Estimated value이면 지출접근법 후보 우선 사용 {payloads['gdp']['summary'].get('한국추정값_지출접근우선셀수', 0)}셀"],
        ["GDP 보강 표시", "지출접근법으로 보강한 셀은 보라색 채우기"],
        ["비확정값 표시", "관측상태가 A가 아닌 셀은 메모만 작성. GDP 지출접근 보강 셀은 보라색 채우기"],
        ["FRED 원천", payloads["fred"]["summary"]["원천파일"]],
        ["비금융자산 원천", payloads["table9b"]["summary"]["원천파일"]],
        ["금융순자산 원천", payloads["financial"]["summary"]["원천파일"]],
        ["GDP 원천", payloads["gdp"]["summary"]["원천파일"]],
    ]
    append_rows(ws, rows)
    format_sheet(ws, freeze_cols=1)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def build_workbook(payloads: dict[str, Any]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_fred_sheet(wb, payloads["fred"])
    table9b_cached_values = build_table9b_sheet(wb, payloads["table9b"])
    build_financial_sheet(wb, payloads["financial"])
    build_gdp_sheet(wb, payloads["gdp"])
    build_info_sheet(wb, payloads)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(FINAL_WORKBOOK)
    add_formula_cached_values(FINAL_WORKBOOK, "비금융자산_피벗", table9b_cached_values)


def verify_workbook() -> dict[str, int | str]:
    wb = load_workbook(FINAL_WORKBOOK)
    comments = purple = yellow = korea_rows = korea_cells = abs_rows = abs_cells = fred_us_rows = fred_us_cells = fred_us_formula_cells = 0
    country_columns = {"비금융자산_피벗": 3, "금융순자산_피벗": 1, "GDP_OECD회원국": 1}
    for ws in wb.worksheets:
        country_col = country_columns.get(ws.title)
        for row in ws.iter_rows():
            is_korea_row = bool(country_col and row[0].row > 1 and ws.cell(row=row[0].row, column=country_col).value == "Korea")
            is_abs_row = ws.title == "금융순자산_피벗" and row[0].row > 1 and ws.cell(row=row[0].row, column=2).value == "AUS_ABS"
            is_fred_us_row = (
                ws.title == "비금융자산_피벗"
                and row[0].row > 1
                and ws.cell(row=row[0].row, column=3).value == "United States"
                and ws.cell(row=row[0].row, column=2).value
                in {"Total non-financial assets, net", "Real estate at market value (FRED 3-series sum)"}
            )
            if is_korea_row:
                korea_rows += 1
            if is_abs_row:
                abs_rows += 1
            if is_fred_us_row:
                fred_us_rows += 1
            for cell in row:
                if cell.comment:
                    comments += 1
                fill = cell.fill.fgColor.rgb
                if fill == "00D9D2E9":
                    purple += 1
                if fill == "00FFF2CC":
                    yellow += 1
                if is_korea_row and fill == "00E2F0D9":
                    font_color = cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else ""
                    if font_color == "000000FF":
                        korea_cells += 1
                if is_abs_row and fill == "00EADCF8":
                    abs_cells += 1
                if is_fred_us_row and fill == "00F4CCCC":
                    fred_us_cells += 1
                if is_fred_us_row and cell.column >= 5 and isinstance(cell.value, str) and cell.value.startswith("=SUM("):
                    fred_us_formula_cells += 1
    data_only_wb = load_workbook(FINAL_WORKBOOK, data_only=True, read_only=True)
    data_only_ws = data_only_wb["비금융자산_피벗"]
    fred_us_cached_value_cells = 0
    for row_no in range(2, data_only_ws.max_row + 1):
        is_fred_us_row = (
            data_only_ws.cell(row=row_no, column=3).value == "United States"
            and data_only_ws.cell(row=row_no, column=2).value
            in {"Total non-financial assets, net", "Real estate at market value (FRED 3-series sum)"}
        )
        if not is_fred_us_row:
            continue
        for col_no in range(5, data_only_ws.max_column + 1):
            if isinstance(data_only_ws.cell(row=row_no, column=col_no).value, (int, float)):
                fred_us_cached_value_cells += 1
    data_only_wb.close()
    return {
        "sheet_count": len(wb.sheetnames),
        "comments": comments,
        "purple_fills": purple,
        "yellow_fills": yellow,
        "korea_highlight_rows": korea_rows,
        "korea_highlight_cells": korea_cells,
        "abs_520410_highlight_rows": abs_rows,
        "abs_520410_highlight_cells": abs_cells,
        "fred_us_derived_highlight_rows": fred_us_rows,
        "fred_us_derived_highlight_cells": fred_us_cells,
        "fred_us_derived_formula_cells": fred_us_formula_cells,
        "fred_us_derived_cached_value_cells": fred_us_cached_value_cells,
        "file_size": FINAL_WORKBOOK.stat().st_size,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect sources and build the requested pivot workbook.")
    parser.add_argument("--skip-normalizers", action="store_true", help="Use existing normalized CSVs.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    log_rows: list[dict[str, str]] = []
    if not args.skip_normalizers:
        for script in NORMALIZER_SCRIPTS:
            log_rows.append(run_step(f"정규화:{script.name}", [sys.executable, str(script)]))
    for script in PAYLOAD_SCRIPTS:
        log_rows.append(run_step(f"payload:{script.name}", [sys.executable, str(script)]))
    payloads = load_payloads()
    started = dt.datetime.now().astimezone()
    build_workbook(payloads)
    source_archive_file_count = create_source_archive()
    verification = verify_workbook()
    verification["source_archive_files"] = source_archive_file_count
    verification["source_archive_size"] = SOURCE_ARCHIVE.stat().st_size
    finished = dt.datetime.now().astimezone()
    log_rows.append(
        {
            "단계": "workbook:openpyxl",
            "명령": str(pathlib.Path(__file__).name),
            "시작일시": started.isoformat(timespec="seconds"),
            "종료일시": finished.isoformat(timespec="seconds"),
            "상태": "완료",
            "종료코드": "0",
            "표준출력": json.dumps(verification, ensure_ascii=False),
            "표준오류": "",
        }
    )
    write_csv(BUILD_LOG, ["단계", "명령", "시작일시", "종료일시", "상태", "종료코드", "표준출력", "표준오류"], log_rows)
    print(json.dumps(verification, ensure_ascii=False))
    print(f"workbook={FINAL_WORKBOOK}")
    print(f"log={BUILD_LOG}")
    print(f"source_archive={SOURCE_ARCHIVE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
